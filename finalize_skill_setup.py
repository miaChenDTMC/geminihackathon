import os
import shutil
import openpyxl
import subprocess
from pathlib import Path

def setup():
    skill_dir = Path("AI Act skills packages/AI Act package/downstream-notifier")
    template_dir = skill_dir / "templates"
    example_dir = skill_dir / "examples"
    
    # 1. Ensure directories exist
    template_dir.mkdir(exist_ok=True)
    example_dir.mkdir(exist_ok=True)
    
    # 2. Copy master templates
    shutil.copy("Output/EU_AI_Act_Annex_XII_Downstream_Provider_Template.xlsx", template_dir / "Annex_XII_Provider_Template.xlsx")
    shutil.copy("Output/EU_AI_Act_Article_13_Compliance_Template.xlsx", template_dir / "Article_13_Compliance_Template.xlsx")
    print("✓ Master templates preserved in skill folder.")
    
    # 3. Create 3 example providers
    providers = [
        ('Alpha_Tech', 'Alpha Tech Solutions', 'P001'),
        ('Beta_Global', 'Beta Global Systems', 'P002'),
        ('Gamma_AI', 'Gamma AI Services', 'P003')
    ]
    
    master_annex = template_dir / "Annex_XII_Provider_Template.xlsx"
    for file_prefix, name, pid in providers:
        # We use openpyxl carefully to avoid merged cell issues
        wb = openpyxl.load_workbook(master_annex)
        ws = wb['Downstream Provider Info']
        # Based on inspection: Row 2 Col 2 is Name, Row 3 Col 2 is ID (if index 1 starts at header)
        # Let's check coordinates again. 
        # Row 1 is header: DOWNSTREAM PROVIDER INFORMATION | Unnamed: 1
        # Row 2 is "Provider Name" | [Empty]
        # Row 3 is "Provider ID" | [Empty]
        ws.cell(row=2, column=2).value = name
        ws.cell(row=3, column=2).value = pid
        
        out_path = example_dir / f"{file_prefix}_Provider.xlsx"
        wb.save(out_path)
        print(f"✓ Created example: {out_path.name}")

    # 4. Setup mock_repo inside examples
    repo_dir = example_dir / "mock_repo"
    repo_dir.mkdir(exist_ok=True)
    os.chdir(repo_dir)
    try:
        if not (repo_dir / ".git").exists():
            subprocess.run(["git", "init"], check=True, capture_output=True)
        
        (repo_dir / "main.py").write_text("print('Hello World')", encoding='utf-8')
        subprocess.run(["git", "add", "."], check=True)
        subprocess.run(["git", "config", "user.email", "tester@example.com"], check=True)
        subprocess.run(["git", "config", "user.name", "Tester"], check=True)
        subprocess.run(["git", "commit", "-m", "Initial commit"], capture_output=True)
        
        # Make a change
        (repo_dir / "main.py").write_text("print('Hello World Updated •')", encoding='utf-8')
        print("✓ example_repo initialized with pending changes.")
    except Exception as e:
        print(f"⚠ git failed in mock_repo: {e}")

if __name__ == "__main__":
    setup()
