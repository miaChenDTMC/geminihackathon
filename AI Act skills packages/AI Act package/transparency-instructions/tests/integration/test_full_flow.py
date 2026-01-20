import pytest
import os
import sys
import shutil
from pathlib import Path
from openpyxl import load_workbook

# Add scripts to path
SCRIPT_DIR = Path(__file__).resolve().parent.parent / "scripts"
sys.path.append(str(SCRIPT_DIR))

import scanner
import excel_injector

# Mock GenAI to avoid API calls during test
class MockResponse:
    def __init__(self, text):
        self.text = text

class MockModel:
    def generate_content(self, model, contents, config=None):
        return MockResponse('{"provider_name": "TestProvider", "13.3.b.i.1": "Test Purpose", "expected_lifetime": "10 years"}')

class MockClient:
    def __init__(self, api_key):
        self.models = MockModel()

def test_scanner_with_fake_repo():
    fake_repo = Path(__file__).resolve().parent.parent / "examples" / "fake_repo"
    files = scanner.scan_codebase(fake_repo)
    
    assert len(files) >= 1
    filenames = [f['path'] for f in files]
    assert "main.py" in filenames or "main.py" in [f.split('/')[-1] for f in filenames]
    assert "README.md" in filenames or "README.md" in [f.split('/')[-1] for f in filenames]

def test_excel_injection_logic(tmp_path):
    # Create dummy excel
    wb = load_workbook(Path(__file__).resolve().parent.parent / "templates" / "Transparency_Instructions_Art13.xlsx")
    # Just use the real template path for reading, save to tmp
    temp_template = tmp_path / "temp_template.xlsx"
    wb.save(temp_template)
    
    output_path = tmp_path / "output.xlsx"
    
    data = {
        "Provider Name": "IntegrationTestProvider",
        "Expected Lifetime": "99 Years"
    }
    
    success = excel_injector.inject_content(str(temp_template), str(output_path), data)
    assert success is True
    assert output_path.exists()
    
    # Verify content
    wb_out = load_workbook(output_path)
    found_provider = False
    for sheet in wb_out.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == "IntegrationTestProvider":
                    found_provider = True
    
    # Note: Depending on where "Provider Name" is in the template, this might vary.
    # But based on our logic, it should be injected.
    assert found_provider
