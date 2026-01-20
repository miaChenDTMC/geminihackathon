"""
Excel Injector for Transparency Instructions
============================================
Handles loading CSV templates, injecting content, and merging into a multi-sheet Excel.
"""

import pandas as pd
import os
from pathlib import Path
from typing import Dict, Any, List

def generate_excel_from_csvs(template_dir: str, output_path: str, data_map: Dict[str, Any]) -> bool:
    """
    Reads 5 specific CSVs, fills them, and creates a combined Excel file.
    """
    csv_files = [
        "Transparency_Instructions_Art13_1. Provider Identity.csv",
        "Transparency_Instructions_Art13_2. Characteristics.csv",
        "Transparency_Instructions_Art13_3. Risks & Limitations.csv",
        "Transparency_Instructions_Art13_4. Oversight & Maintenance.csv",
        "Transparency_Instructions_Art13_5. Logging.csv"
    ]
    
    sheet_names = [
        "1. Provider Identity",
        "2. Characteristics",
        "3. Risks & Limitations",
        "4. Oversight & Maintenance",
        "5. Logging"
    ]
    
    template_path = Path(template_dir)
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for csv_file, sheet_name in zip(csv_files, sheet_names):
                full_csv_path = template_path / csv_file
                if not full_csv_path.exists():
                    print(f"Error: Missing CSV template: {full_csv_path}")
                    return False
                
                # Read CSV
                # Keep default header=0 (first line is header)
                try:
                    df = pd.read_csv(full_csv_path, encoding='utf-8')
                except UnicodeDecodeError:
                    df = pd.read_csv(full_csv_path, encoding='latin1')
                
                # Column 0 is Label, Column 1 is Input Target
                # Normalize column names just in case, but assume index 0/1 are valid
                label_col = df.columns[0]
                input_col = df.columns[1]
                
                print(f"Processing {sheet_name}...")
                
                # Iterate and Fill
                for idx, row in df.iterrows():
                    label = str(row[label_col]).strip()
                    if not label or label == "nan": continue
                    
                    # Direct Match
                    match_val = data_map.get(label)
                    
                    if match_val:
                        print(f"  MATCH: '{label}'")
                        
                        if isinstance(match_val, dict):
                            # Handle Multi-Column Injection
                            # match_val = {"Description/Metric": "...", "Validation Reference": "..."}
                            for target_col_name, target_val in match_val.items():
                                # Find column index by name (case-insensitive?)
                                # df.columns contains the headers
                                if target_col_name in df.columns and target_val:
                                    df.at[idx, target_col_name] = target_val
                                    print(f"    -> Writing to '{target_col_name}'")
                        else:
                            # Handle Single Value (Legacy/Default) -> Write to 2nd Column
                            print(f"    -> Writing to '{input_col}'")
                            df.at[idx, input_col] = match_val
                    else:
                         # Attempt fuzzy/substring matches if needed
                         pass
                
                # Validate DataFrame isn't empty
                # Write to Excel Sheet
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"  > Added sheet: {sheet_name}")
                
        # --- Formatting Pass using openpyxl ---
        from openpyxl import load_workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

        wb = load_workbook(output_path)
        
        # User Styles
        thin = Side(border_style="thin", color="000000")
        all_borders = Border(top=thin, left=thin, right=thin, bottom=thin)
        black_font = Font(name='Calibri', size=11, color="000000") 
        wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
        
        # FIX: User requested "White", not "No Fill" (which can appear gray)
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for ws in wb.worksheets:
            # 1. Adjust styles and widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                
                for cell in col:
                    cell.font = black_font
                    cell.border = all_borders
                    cell.alignment = wrap_align
                    cell.fill = white_fill
                    
                    try: 
                        val_len = len(str(cell.value))
                        if val_len > max_length: max_length = val_len
                    except: pass
                
                adjusted_width = (max_length + 2)
                if adjusted_width > 60: adjusted_width = 60
                if adjusted_width < 15: adjusted_width = 15
                ws.column_dimensions[column].width = adjusted_width

            # 2. Add Table
            if ws.max_row > 1:
                ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                # Safe table name
                safe_name = "".join([c for c in ws.title if c.isalnum()])
                tab_name = f"Tbl_{safe_name}"
                
                tab = Table(displayName=tab_name, ref=ref)
                style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                tab.tableStyleInfo = style
                ws.add_table(tab)
        
        wb.save(output_path)
        print(f"Successfully saved and formatted Excel to: {output_path}")
        return True

    except Exception as e:
        print(f"Error generating Excel from CSVs: {e}")
        return False
