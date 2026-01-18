"""Verify which cells are filled vs empty in the Zenith output."""
import openpyxl
from pathlib import Path

lb = sorted(Path('Output').glob('Batch_*'))[-1]
print(f"Latest Batch: {lb}")

for fname in lb.glob('*.xlsx'):
    print(f"\n=== {fname.name} ===")
    wb = openpyxl.load_workbook(fname, data_only=True)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        filled = 0
        empty = 0
        empty_labels = []
        
        for row in sheet.iter_rows(min_row=1, max_row=60):
            label = str(row[0].value)[:30] if row[0].value else None
            for cell in row[1:5]:  # Check columns B-E
                if cell.value and str(cell.value).strip():
                    filled += 1
                else:
                    empty += 1
                    if label:
                        empty_labels.append(f"R{cell.row}:{label}")
        
        print(f"  {sheet_name}: Filled={filled}, Empty={empty}")
        if empty_labels[:10]:
            print(f"    Empty at: {empty_labels[:10]}")
