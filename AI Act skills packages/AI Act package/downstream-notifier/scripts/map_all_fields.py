"""
Comprehensive Field Mapper for EU AI Act Compliance Templates.
Maps ALL empty fields that need to be filled by the LLM.
"""
import openpyxl
from pathlib import Path

def map_template(template_path):
    wb = openpyxl.load_workbook(template_path, data_only=True)
    field_map = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        field_map[sheet_name] = []
        
        for row_idx in range(1, min(100, sheet.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(10, sheet.max_column + 1)):
                cell = sheet.cell(row=row_idx, column=col_idx)
                val = cell.value
                row_data.append({
                    'row': row_idx,
                    'col': col_idx,
                    'value': str(val)[:50] if val else '[EMPTY]',
                    'is_empty': val is None or str(val).strip() == ''
                })
            field_map[sheet_name].append(row_data)
    
    return field_map

def print_empty_fields(field_map, template_name):
    print(f"\n{'='*60}")
    print(f"TEMPLATE: {template_name}")
    print(f"{'='*60}")
    
    for sheet_name, rows in field_map.items():
        print(f"\n--- {sheet_name} ---")
        for row in rows:
            # Find rows with labels in Column A/B and empty input columns
            label = next((c['value'] for c in row if c['col'] <= 2 and c['value'] != '[EMPTY]'), None)
            empty_cols = [c['col'] for c in row if c['is_empty'] and c['col'] > 1]
            if label and empty_cols:
                print(f"  R{row[0]['row']}: {label[:40]} -> Empty Cols: {empty_cols}")

# Map both templates
templates_dir = Path('templates')
for tmpl in templates_dir.glob('*.xlsx'):
    if 'backup' not in str(tmpl).lower() and '~$' not in tmpl.name:
        fm = map_template(tmpl)
        print_empty_fields(fm, tmpl.name)
