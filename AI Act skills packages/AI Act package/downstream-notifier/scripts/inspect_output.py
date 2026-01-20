"""Detailed output inspection."""
import openpyxl
from pathlib import Path

lb = sorted(Path('Output').glob('Batch_*'))[-1]
print(f"Latest Batch: {lb}")

for f in lb.glob('Article_13_Compliance_*.xlsx'):
    print(f"\n{'='*60}")
    print(f"FILE: {f.name}")
    print('='*60)
    wb = openpyxl.load_workbook(f)
    
    for sn in wb.sheetnames:
        sheet = wb[sn]
        print(f"\n--- SHEET: {sn} ---")
        
        # Print rows with their content
        for row_idx in range(1, min(35, sheet.max_row + 1)):
            row_data = []
            has_content = False
            for col_idx in range(1, min(7, sheet.max_column + 1)):
                val = sheet.cell(row=row_idx, column=col_idx).value
                if val:
                    has_content = True
                val_str = str(val)[:40] if val else '<EMPTY>'
                row_data.append(f"C{col_idx}:{val_str}")
            
            # Only print rows with some content or specific rows
            if has_content or row_idx <= 6:
                print(f"R{row_idx}: {' | '.join(row_data)}")
    break  # Just check first file
