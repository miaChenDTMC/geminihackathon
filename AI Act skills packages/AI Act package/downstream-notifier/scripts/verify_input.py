"""Verify fake input file has all cells filled."""
import openpyxl
from pathlib import Path

ef = Path('examples/Fake_Annex_XII_Input.xlsx')
wb = openpyxl.load_workbook(ef)

for sn in wb.sheetnames:
    sheet = wb[sn]
    total = 0
    empty = 0
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            total += 1
            if not cell.value or str(cell.value).strip() == '':
                empty += 1
    print(f"{sn}: Total={total}, Empty={empty}, Filled={total-empty}")
