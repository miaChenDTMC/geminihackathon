"""Verify Article 13 output file has 0 empty cells."""
import openpyxl
from pathlib import Path

# Find latest batch
lb = sorted(Path('Output').glob('Batch_*'))[-1]
print(f"Latest Batch: {lb}")

# Check Article 13 output for one provider
for f in lb.glob('Article_13_Compliance_*.xlsx'):
    print(f"\n=== {f.name} ===")
    wb = openpyxl.load_workbook(f)
    
    for sn in wb.sheetnames:
        sheet = wb[sn]
        total = 0
        empty = 0
        unchecked = 0
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                total += 1
                val = str(cell.value) if cell.value else ""
                if not val.strip():
                    empty += 1
                if "☐" in val and "☑" not in val:
                    unchecked += 1
        print(f"  {sn}: Total={total}, Empty={empty}, Unchecked={unchecked}")
    break  # Just check first file
