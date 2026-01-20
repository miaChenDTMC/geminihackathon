"""
Generate Fake Annex XII Input Excel.
Scans ALL worksheets and ALL cells in the template.
Fills every empty cell with appropriate fake data or ticks checkboxes.
"""
import openpyxl
from pathlib import Path
from datetime import datetime
import re

# Fake data pool for realistic values
# Comprehensive Fake Data Pool for realistic Annex XII simulation
FAKE_DATA = {
    "provider": {
        "name": "Zenith AI Systems Corp",
        "id": "GPAI-2026-ZEN-001",
        "email": "compliance@zenith-ai.com",
        "phone": "+1-555-0199",
        "address": "123 Innovation Drive, Silicon Valley, CA 94025",
        "representative": "Dr. Jane Smith, Chief Compliance Officer",
        "website": "https://zenith-ai.com",
    },
    "model": {
        "name": "ZenithGPT-4 Enterprise",
        "version": "v4.2.1",
        "release_date": "2026-01-15",
        "architecture": "70B Parameter Transformer, Decoder-only, sparsely activated Mixture-of-Experts (MoE)",
        "parameters": "70 Billion (active parameters approx 12B/token)",
        "context_window": "128,000 tokens (approx 100 pages of text)",
        "license": "Enterprise Commercial License v2.0 (Proprietary)",
    },
    "technical": {
        "input_modality": "Text (UTF-8, Markdown), Structured Data (JSON, XML), Python Code",
        "output_modality": "Natural Language Generation, Code Completion, Semantic Summarization",
        "hardware": "Minimum: NVIDIA A10G (24GB VRAM) or equivalent. Recommended: A100 (80GB). Storage: 100GB SSD.",
        "software": "Docker Container (v24.0+), Python 3.11+, CUDA 12.1 drivers. Orchestration via Kubernetes supported.",
        "api": "RESTful API (OpenAPI 3.1 compliant), gRPC support for low-latency streaming. Rate limits: 50 RPM.",
        "rate_limit": "Standard Tier: 1000 requests/minute. Enterprise: Custom limits available.",
        "energy": "Estimated 0.4 kWh per 1M tokens processed. Carbon offset certificates available upon request.",
    },
    "lifecycle": {
        "lifetime": "Supported operational lifetime: 36 months from version release. LTS support available until 2030.",
        "maintenance": "Quarterly security patches. Monthly model fine-tuning updates for alignment. Weekly downtime window: Sunday 02:00 UTC.",
        "updates": "OTA (Over-the-Air) container image updates via private registry. Automated rollback enabled for failed deployments.",
        "validation": "Red-teaming performed by external auditors (CertAI Ltd). F1-score > 0.92 on medical reasoning benchmarks.",
    },
    "compliance": {
        "intended_tasks": "Enterprise knowledge retrieval, automated customer support, code refactoring, document synthesis.",
        "prohibited_uses": "Biometric identification, social scoring, generation of non-consensual deepfakes, autonomous weapons targeting.",
        "acceptable_use": "Internal business operations, software development assistance, academic research analysis.",
        "training_data": "dataset-metadata-v4.json (attached). Includes CommonCrawl (filtered), StackOverflow, Wikipedia, and licensed medical journals.",
        "risk_level": "Limited Risk (under Article 6 classification) - Transparency obligations apply.",
        "gdpr": "GDPR compliant data filtering pipeline. 'Right to be Forgotten' supported via unlearning requests.",
    }
}

def get_fake_value(cell_label):
    """Generate fake value based on the cell's label with improved matching."""
    label = str(cell_label).lower() if cell_label else ""
    
    # Provider info
    if "provider name" in label or "identity of provider" in label: return FAKE_DATA["provider"]["name"]
    if "email" in label or "contact" in label: return FAKE_DATA["provider"]["email"]
    if "phone" in label: return FAKE_DATA["provider"]["phone"]
    if "address" in label: return FAKE_DATA["provider"]["address"]
    if "representative" in label: return FAKE_DATA["provider"]["representative"]
    if "registration" in label or "provider id" in label: return FAKE_DATA["provider"]["id"]
    if "website" in label or "url" in label: return FAKE_DATA["provider"]["website"]
    
    # Model info
    if "model name" in label or "system name" in label: return FAKE_DATA["model"]["name"]
    if "version" in label: return FAKE_DATA["model"]["version"]
    if "release date" in label: return FAKE_DATA["model"]["release_date"]
    if "architecture" in label: return FAKE_DATA["model"]["architecture"]
    if "parameter" in label: return FAKE_DATA["model"]["parameters"]
    if "context" in label or "token" in label: return FAKE_DATA["model"]["context_window"]
    if "license" in label: return FAKE_DATA["model"]["license"]
    if "energy" in label or "power" in label or "consumption" in label: return FAKE_DATA["technical"]["energy"]
    
    # Technical
    if "input" in label and "modal" in label: return FAKE_DATA["technical"]["input_modality"]
    if "output" in label and "modal" in label: return FAKE_DATA["technical"]["output_modality"]
    if "hardware" in label or "compute" in label or "gpu" in label: return FAKE_DATA["technical"]["hardware"]
    if "software" in label or "dependency" in label: return FAKE_DATA["technical"]["software"]
    if "api" in label or "interface" in label: return FAKE_DATA["technical"]["api"]
    
    # Lifecycle (New detailed fields)
    if "lifetime" in label or "life-cycle" in label: return FAKE_DATA["lifecycle"]["lifetime"]
    if "maintain" in label or "maintenance" in label: return FAKE_DATA["lifecycle"]["maintenance"]
    if "update" in label or "frequency" in label: return FAKE_DATA["lifecycle"]["updates"]
    if "validat" in label or "test" in label or "accura" in label: return FAKE_DATA["lifecycle"]["validation"]
    
    # Compliance
    if "intended" in label or "task" in label or "purpose" in label: return FAKE_DATA["compliance"]["intended_tasks"]
    if "prohibited" in label or "restriction" in label: return FAKE_DATA["compliance"]["prohibited_uses"]
    if "acceptable" in label: return FAKE_DATA["compliance"]["acceptable_use"]
    if "training" in label or "data" in label: return FAKE_DATA["compliance"]["training_data"]
    if "risk" in label: return FAKE_DATA["compliance"]["risk_level"]
    if "gdpr" in label or "privacy" in label: return FAKE_DATA["compliance"]["gdpr"]
    
    # Dates
    if "date" in label: return datetime.now().strftime("%Y-%m-%d")
    
    # Fallback for "Review and confirm" fields if no specific match
    if "review" in label: return "Technical specifications reviewed and confirmed by CCO."
    
    # Final generic fallback with more professional tone
    return f"Specific details for '{cell_label[:20]}...' available in attached Technical Documentation Annex A."

def is_checkbox_cell(value):
    """Check if cell contains a checkbox pattern."""
    val = str(value) if value else ""
    return "☐" in val or "☑" in val or "checkbox" in val.lower()

def tick_checkbox(value):
    """Convert unchecked to checked checkbox."""
    val = str(value) if value else ""
    # Replace unchecked with checked
    return val.replace("☐ Yes / ☐ No", "☑ Yes / ☐ No").replace("☐ Complete", "☑ Complete").replace("☐", "☑")

def generate_fake_input(template_path, output_path):
    """Generate fully populated fake Annex XII input file."""
    wb = openpyxl.load_workbook(template_path)
    
    stats = {"filled": 0, "ticked": 0, "skipped": 0}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Processing: {sheet_name} ---")
        
        for row_idx in range(1, sheet.max_row + 1):
            # Get label from column A or B
            label_cell = sheet.cell(row=row_idx, column=1)
            label = str(label_cell.value) if label_cell.value else ""
            
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                try:
                    current_val = cell.value
                    
                    # Skip if already has meaningful content
                    if current_val and str(current_val).strip() and not is_checkbox_cell(current_val):
                        stats["skipped"] += 1
                        continue
                    
                    # Handle checkboxes
                    if is_checkbox_cell(current_val):
                        cell.value = tick_checkbox(current_val)
                        stats["ticked"] += 1
                        print(f"  Ticked: R{row_idx}C{col_idx}")
                        continue
                    
                    # Fill empty cells (column B onwards typically input fields)
                    if col_idx > 1 and (not current_val or str(current_val).strip() == ""):
                        fake_val = get_fake_value(label)
                        cell.value = fake_val
                        stats["filled"] += 1
                        print(f"  Filled: R{row_idx}C{col_idx} <- {fake_val[:30]}...")
                        
                except AttributeError:
                    # Skip merged cells
                    pass
    
    # Save
    wb.save(output_path)
    wb.close()
    
    print(f"\n=== Summary ===")
    print(f"Cells filled: {stats['filled']}")
    print(f"Checkboxes ticked: {stats['ticked']}")
    print(f"Cells skipped (already had content): {stats['skipped']}")
    print(f"Output saved to: {output_path}")
    
    return output_path

if __name__ == "__main__":
    template = Path("templates/Annex_XII_Provider_Template.xlsx")
    output = Path("examples/Fake_Annex_XII_Input.xlsx")
    output.parent.mkdir(exist_ok=True)
    generate_fake_input(template, output)
