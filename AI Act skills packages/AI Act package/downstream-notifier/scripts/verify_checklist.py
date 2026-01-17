import openpyxl
from pathlib import Path

lb = sorted(Path('Output').glob('Batch_*'))[-1]
ef = next(lb.glob('Annex_XII_Provider_Z-9000-X.xlsx'))
wb = openpyxl.load_workbook(ef, data_only=True)
print(f"Sheets: {wb.sheetnames}")

# Check all sheets for checklist marks and metadata
for sn in wb.sheetnames:
    print(f"\n=== {sn} ===")
    sheet = wb[sn]
    for r in range(1, min(35, sheet.max_row + 1)):
        row_vals = [str(c.value)[:40] if c.value else '' for c in sheet[r][:6]]
        # Show rows with XII references, Provider labels, or checkbox symbols
        if any('XII' in v or 'Provider' in v or '☑' in v or 'Downstream' in v for v in row_vals):
            print(f"R{r}: {row_vals}")
