"""Find all empty cells that SHOULD have values."""
import openpyxl
from pathlib import Path

lb = sorted(Path('Output').glob('Batch_*'))[-1]
print(f"Latest Batch: {lb}")

# Row limits per sheet
row_limits = {
    "Instructions for Use": 62,
    "Compliance Checklist": 30,
    "Document Metadata": 16,
    "GPAI Model Documentation": 56,
    "Downstream Provider Info": 15,
}

for f in lb.glob('Article_13_Compliance_*.xlsx'):
    print(f"\nFILE: {f.name}")
    wb = openpyxl.load_workbook(f)
    
    for sn in wb.sheetnames:
        sheet = wb[sn]
        max_row = row_limits.get(sn, 20)
        print(f"\n--- {sn} (rows 5 to {max_row}) ---")
        
        empty_cells = []
        for row_idx in range(5, max_row + 1):
            col_a = str(sheet.cell(row=row_idx, column=1).value or "").strip()
            if not col_a:  # Skip rows with no label
                continue
                
            for col_idx in range(2, 7):
                val = sheet.cell(row=row_idx, column=col_idx).value
                if not val or str(val).strip() == "":
                    header = str(sheet.cell(row=4, column=col_idx).value or "")[:20]
                    empty_cells.append(f"R{row_idx}C{col_idx} ({col_a} / {header})")
        
        if empty_cells:
            print(f"Empty cells: {len(empty_cells)}")
            for ec in empty_cells[:10]:
                print(f"  - {ec}")
        else:
            print("All cells filled!")
    break
