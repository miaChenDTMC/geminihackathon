#!/usr/bin/env python3
"""
Download GitHub repositories from AI_Risk_Assessment_Guide.xlsx Risk Tools sheet
into Risks packages folder following category logic.
"""

import openpyxl
import os
import subprocess
from urllib.parse import urlparse
import json
from datetime import datetime

# Category mapping from Excel to folder names
CATEGORY_MAPPING = {
    'Legal': 'Legal',
    'Cybersecurity': 'Cybersecurity',
    'Environment': 'Environment',
    'Technical': 'Technical',
    'Trust': 'Trust',
    'Fundamental Rights': 'Fundamental Rights',
    'Privacy': 'EU AI Act Compliance',  # Privacy tools go to EU AI Act Compliance
    'Societal': 'Societal',
    'Third-Party': 'Third-Party',
    'Business': 'Societal',  # Business tools go to Societal
    'Health & Safety': 'Health & Safety',
    'GDPR Compliance': 'EU AI Act Compliance'  # GDPR tools go to EU AI Act Compliance
}

def is_github_url(url):
    """Check if URL is a GitHub repository URL."""
    if not url or not isinstance(url, str):
        return False
    return 'github.com' in url.lower() and url.startswith('http')

def get_repo_name(github_url):
    """Extract repository name from GitHub URL."""
    # Parse URL like https://github.com/user/repo
    parts = github_url.rstrip('/').split('/')
    if len(parts) >= 2:
        return parts[-1]  # Return repo name
    return None

def clone_repository(github_url, target_dir, tool_name):
    """Clone a GitHub repository to target directory."""
    repo_name = get_repo_name(github_url)
    if not repo_name:
        return False, "Could not parse repository name"

    # Create target directory if it doesn't exist
    os.makedirs(target_dir, exist_ok=True)

    # Full path for the cloned repo
    repo_path = os.path.join(target_dir, repo_name)

    # Check if already exists
    if os.path.exists(repo_path):
        return True, f"Already exists: {repo_path}"

    # Clone the repository
    try:
        print(f"  Cloning {github_url} to {repo_path}...")
        result = subprocess.run(
            ['git', 'clone', '--depth', '1', github_url, repo_path],
            capture_output=True,
            text=True,
            timeout=300  # 5 minute timeout
        )

        if result.returncode == 0:
            return True, f"Successfully cloned to {repo_path}"
        else:
            return False, f"Git clone failed: {result.stderr}"
    except subprocess.TimeoutExpired:
        return False, "Clone timeout (5 minutes)"
    except Exception as e:
        return False, f"Error: {str(e)}"

def main():
    """Main function to process Excel and download repositories."""

    # Base paths
    base_dir = '/Users/miachen/Desktop/Mia/DT Master/Tech/Hackthon/geminihackathon'
    excel_file = os.path.join(base_dir, 'AI_Risk_Assessment_Guide.xlsx')
    risks_packages_dir = os.path.join(base_dir, 'Risks packages')

    # Load workbook
    print("Loading AI_Risk_Assessment_Guide.xlsx...")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Risk Tools']

    # Parse header
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # Find relevant column indices
    category_idx = headers.index('Category')
    tool_name_idx = headers.index('Tool Name')
    endpoint_idx = headers.index('Endpoint/GitHub')
    to_download_idx = headers.index('To Download')

    # Statistics
    stats = {
        'total_tools': 0,
        'github_tools': 0,
        'downloaded': 0,
        'skipped': 0,
        'failed': 0,
        'already_exists': 0,
        'by_category': {}
    }

    # Detailed log
    download_log = []

    print("\nProcessing tools from Risk Tools sheet...\n")

    # Process each row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[category_idx]:
            continue

        stats['total_tools'] += 1

        category = row[category_idx]
        tool_name = row[tool_name_idx]
        github_url = row[endpoint_idx]
        to_download = row[to_download_idx]

        # Initialize category stats
        if category not in stats['by_category']:
            stats['by_category'][category] = {
                'total': 0,
                'downloaded': 0,
                'skipped': 0,
                'failed': 0
            }

        stats['by_category'][category]['total'] += 1

        # Check if it's a GitHub URL and marked for download
        if is_github_url(github_url):
            stats['github_tools'] += 1

            print(f"[{row_idx}] {category} - {tool_name}")
            print(f"  URL: {github_url}")
            print(f"  To Download: {to_download}")

            if to_download == 'Yes':
                # Map category to folder
                target_folder = CATEGORY_MAPPING.get(category, category)
                target_dir = os.path.join(risks_packages_dir, target_folder)

                # Clone repository
                success, message = clone_repository(github_url, target_dir, tool_name)

                print(f"  {message}")

                log_entry = {
                    'row': row_idx,
                    'category': category,
                    'tool_name': tool_name,
                    'github_url': github_url,
                    'target_folder': target_folder,
                    'success': success,
                    'message': message
                }
                download_log.append(log_entry)

                if success:
                    if 'Already exists' in message:
                        stats['already_exists'] += 1
                        stats['by_category'][category]['skipped'] += 1
                    else:
                        stats['downloaded'] += 1
                        stats['by_category'][category]['downloaded'] += 1
                else:
                    stats['failed'] += 1
                    stats['by_category'][category]['failed'] += 1

                print()
            else:
                print(f"  Skipped (To Download = {to_download})\n")
                stats['skipped'] += 1
                stats['by_category'][category]['skipped'] += 1

    # Print summary
    print("\n" + "="*80)
    print("DOWNLOAD SUMMARY")
    print("="*80)
    print(f"Total tools in sheet: {stats['total_tools']}")
    print(f"Tools with GitHub URLs: {stats['github_tools']}")
    print(f"Successfully downloaded: {stats['downloaded']}")
    print(f"Already existed: {stats['already_exists']}")
    print(f"Skipped: {stats['skipped']}")
    print(f"Failed: {stats['failed']}")

    print("\n" + "-"*80)
    print("BY CATEGORY:")
    print("-"*80)
    for category, cat_stats in sorted(stats['by_category'].items()):
        print(f"\n{category}:")
        print(f"  Total: {cat_stats['total']}")
        print(f"  Downloaded: {cat_stats['downloaded']}")
        print(f"  Skipped: {cat_stats['skipped']}")
        print(f"  Failed: {cat_stats['failed']}")

    # Save detailed log
    log_file = os.path.join(risks_packages_dir, 'DOWNLOAD_LOG.json')
    with open(log_file, 'w') as f:
        json.dump({
            'timestamp': datetime.now().isoformat(),
            'statistics': stats,
            'downloads': download_log
        }, f, indent=2)

    print(f"\n\nDetailed log saved to: {log_file}")

    # Create markdown summary
    md_file = os.path.join(risks_packages_dir, 'DOWNLOAD_SUMMARY_NEW.md')
    with open(md_file, 'w') as f:
        f.write("# Risk Tools Download Summary\n\n")
        f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write("## Statistics\n\n")
        f.write(f"- Total tools in sheet: {stats['total_tools']}\n")
        f.write(f"- Tools with GitHub URLs: {stats['github_tools']}\n")
        f.write(f"- Successfully downloaded: {stats['downloaded']}\n")
        f.write(f"- Already existed: {stats['already_exists']}\n")
        f.write(f"- Skipped: {stats['skipped']}\n")
        f.write(f"- Failed: {stats['failed']}\n\n")

        f.write("## By Category\n\n")
        for category, cat_stats in sorted(stats['by_category'].items()):
            target_folder = CATEGORY_MAPPING.get(category, category)
            f.write(f"### {category} → `{target_folder}`\n\n")
            f.write(f"- Total: {cat_stats['total']}\n")
            f.write(f"- Downloaded: {cat_stats['downloaded']}\n")
            f.write(f"- Skipped: {cat_stats['skipped']}\n")
            f.write(f"- Failed: {cat_stats['failed']}\n\n")

        f.write("## Downloaded Tools\n\n")
        for entry in download_log:
            if entry['success'] and 'Already exists' not in entry['message']:
                f.write(f"- **{entry['tool_name']}** ({entry['category']})\n")
                f.write(f"  - URL: {entry['github_url']}\n")
                f.write(f"  - Location: `Risks packages/{entry['target_folder']}/`\n\n")

        if stats['failed'] > 0:
            f.write("## Failed Downloads\n\n")
            for entry in download_log:
                if not entry['success']:
                    f.write(f"- **{entry['tool_name']}** ({entry['category']})\n")
                    f.write(f"  - URL: {entry['github_url']}\n")
                    f.write(f"  - Error: {entry['message']}\n\n")

    print(f"Markdown summary saved to: {md_file}")

if __name__ == '__main__':
    main()
