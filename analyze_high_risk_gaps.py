#!/usr/bin/env python3
"""
Analyze high-risk AI gaps and create tracking sheet in Excel.
Extracts missing functionality from gap analysis and categorizes by automation potential.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Path to the Excel file
EXCEL_PATH = "/Users/miachen/Library/CloudStorage/OneDrive-DTMasterCarbon/DT Master Mia Personal/5 Tech/AI act/AI_Risk_Assessment_Guide.xlsx"

# Missing functionality data extracted from HIGH_RISK_GAP_ANALYSIS.md
MISSING_FUNCTIONALITY = [
    {
        "category": "Quality Management System (QMS)",
        "functionality": "Quality Policy and Objectives",
        "description": "Formal quality policy signed by executive defining quality objectives, standards, and commitments for AI system",
        "code_automatable": "No",
        "ai_automatable": "Partial",
        "automation_notes": "AI can draft policy templates and suggest objectives based on best practices, but executive approval and organizational context require human input",
        "risk_addressed": "Organizational governance, systematic quality assurance",
        "eu_articles": "Article 17, Annex V.1",
        "effort_hours": 40
    },
    {
        "category": "Quality Management System (QMS)",
        "functionality": "Organizational Structure (RACI Matrix)",
        "description": "Defined roles and responsibilities including AI Product Owner, Quality Manager, DPO, Ethics Committee using RACI framework",
        "code_automatable": "No",
        "ai_automatable": "Partial",
        "automation_notes": "AI can generate RACI template and suggest role definitions, but organizational structure decisions require human management",
        "risk_addressed": "Accountability, clear responsibility assignment, governance structure",
        "eu_articles": "Article 17, Annex V.2",
        "effort_hours": 20,
    },
    {
        "category": "Quality Management System (QMS)",
        "functionality": "Design and Development Procedures",
        "description": "Documented development lifecycle, design reviews, change control procedures, version management",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "AI can generate documentation from code analysis, create design review templates, automate version tracking. Requires integration with git and development tools",
        "risk_addressed": "Development quality, traceability, systematic design approach",
        "eu_articles": "Article 17, Annex V.3",
        "effort_hours": 80,
    },
    {
        "category": "Quality Management System (QMS)",
        "functionality": "Data Management Procedures",
        "description": "Data governance policy, quality checks, version control, data lineage tracking for EU AI Act text and system data",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Can be fully automated with code: data quality checks, version control, checksums, automated validation. AI can enhance with anomaly detection",
        "risk_addressed": "Data quality, data integrity, regulatory compliance",
        "eu_articles": "Article 10, Article 17, Annex V.4",
        "effort_hours": 60,
    },
    {
        "category": "Quality Management System (QMS)",
        "functionality": "Testing and Validation Procedures",
        "description": "Test strategy, automated regression testing, performance testing, security testing, test coverage tracking",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Fully automatable with testing frameworks (pytest, selenium). AI can generate test cases, identify edge cases, optimize test coverage",
        "risk_addressed": "System reliability, accuracy, security validation",
        "eu_articles": "Article 9.6-9.8, Article 15, Annex V.5",
        "effort_hours": 120,
    },
    {
        "category": "Quality Management System (QMS)",
        "functionality": "Post-Market Monitoring System",
        "description": "Continuous monitoring of AI system performance, user feedback collection, incident tracking, performance metrics dashboard",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Fully automatable: telemetry, logging, analytics dashboards, automated alerting. AI can detect anomalies, predict issues, analyze feedback",
        "risk_addressed": "Ongoing safety, performance degradation detection, user satisfaction",
        "eu_articles": "Article 72, Annex V.6",
        "effort_hours": 260,
    },
    {
        "category": "Quality Management System (QMS)",
        "functionality": "Incident Management Procedures",
        "description": "Article 73 serious incident reporting with 15-day timeline, incident classification, root cause analysis, remediation tracking",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "Automation possible for incident detection, classification, notification. AI can assist with root cause analysis, but human oversight required for serious incidents",
        "risk_addressed": "Safety incidents, regulatory reporting, continuous improvement",
        "eu_articles": "Article 73, Annex V.7",
        "effort_hours": 40,
    },
    {
        "category": "Quality Management System (QMS)",
        "functionality": "Documentation and Record-Keeping",
        "description": "Centralized document control system, 10-year retention, audit trail, version control, access logging",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Fully automatable with document management systems, git, database logging. AI can assist with document classification, search, compliance checking",
        "risk_addressed": "Compliance traceability, audit readiness, knowledge management",
        "eu_articles": "Article 11, Article 12, Annex V.8",
        "effort_hours": 60,
    },
    {
        "category": "Quality Management System (QMS)",
        "functionality": "Corrective and Preventive Actions (CAPA)",
        "description": "Non-conformance tracking, root cause analysis, corrective action implementation, effectiveness verification",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "Workflow automation possible. AI can assist with root cause analysis, suggest corrective actions, predict preventive measures",
        "risk_addressed": "Continuous improvement, systematic problem resolution",
        "eu_articles": "Article 17, Annex V.9",
        "effort_hours": 40,
    },
    {
        "category": "Quality Management System (QMS)",
        "functionality": "Management Review and Audit",
        "description": "Quarterly management reviews, annual internal/external audits, audit findings tracking, management oversight",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "Scheduling and reporting can be automated. AI can prepare audit materials, analyze compliance data, identify gaps. Human auditors still required",
        "risk_addressed": "Governance oversight, compliance verification, strategic direction",
        "eu_articles": "Article 17, Annex V.10",
        "effort_hours": 80,
    },
    {
        "category": "Conformity Assessment",
        "functionality": "Internal Control Procedure (Self-Assessment)",
        "description": "Complete technical documentation (Annex IV), compliance verification against Articles 8-15, internal testing, EU Declaration of Conformity",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "AI can generate documentation from code/systems, automate compliance checking, create declaration templates. Human review and signature required",
        "risk_addressed": "Legal compliance, market access, CE marking eligibility",
        "eu_articles": "Article 43, Annex VI",
        "effort_hours": 160,
    },
    {
        "category": "Conformity Assessment",
        "functionality": "Third-Party Assessment by Notified Body",
        "description": "Preparation for notified body review: technical documentation, QMS assessment readiness, testing evidence, surveillance audit preparation",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "AI can organize evidence, generate compliance reports, identify gaps. Actual notified body assessment cannot be automated",
        "risk_addressed": "Independent verification, high-assurance compliance, certification",
        "eu_articles": "Article 43, Annex VII",
        "effort_hours": 80,
    },
    {
        "category": "Regulatory Compliance",
        "functionality": "EU Database Registration",
        "description": "Registration of high-risk AI system in EU database with system details, conformity information, updates when changes occur",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "Form filling can be partially automated. AI can extract system information, prepare registration data. Actual submission requires human oversight",
        "risk_addressed": "Regulatory transparency, market surveillance, public accountability",
        "eu_articles": "Article 71",
        "effort_hours": 40,
    },
    {
        "category": "Post-Market Monitoring",
        "functionality": "Post-Market Monitoring Plan (PMM Plan)",
        "description": "Systematic plan for monitoring: performance metrics, data collection methods, analysis procedures, reporting cadence, alert thresholds",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Fully automatable: metrics collection, dashboard creation, automated analysis, alerting. AI can detect trends, predict issues, optimize monitoring",
        "risk_addressed": "Continuous safety monitoring, performance tracking, early warning",
        "eu_articles": "Article 72",
        "effort_hours": 100,
    },
    {
        "category": "Post-Market Monitoring",
        "functionality": "Performance Metrics Tracking",
        "description": "Track response accuracy (>95%), hallucination rate (<5%), privacy incidents (0), user satisfaction (>4/5), prompt injection attempts, citation accuracy (>99%)",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Fully automatable with telemetry and analytics. AI can enhance with predictive analytics, anomaly detection, automated testing",
        "risk_addressed": "Quality assurance, safety metrics, performance degradation detection",
        "eu_articles": "Article 15, Article 72",
        "effort_hours": 80,
    },
    {
        "category": "Post-Market Monitoring",
        "functionality": "User Feedback Collection and Analysis",
        "description": "Systematic collection of user feedback, sentiment analysis, issue categorization, feedback-driven improvements",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Fully automatable: feedback forms, NLP sentiment analysis, categorization, trend detection. AI excels at analyzing unstructured feedback",
        "risk_addressed": "User satisfaction, usability issues, real-world problem identification",
        "eu_articles": "Article 72",
        "effort_hours": 60,
    },
    {
        "category": "Fundamental Rights",
        "functionality": "Mandatory FRIA - Stakeholder Consultation",
        "description": "Consultation with affected persons, independent experts, civil society organizations; documentation of input received and how it was incorporated",
        "code_automatable": "No",
        "ai_automatable": "Partial",
        "automation_notes": "AI can help schedule consultations, analyze feedback, summarize inputs. Actual consultation requires human interaction and judgment",
        "risk_addressed": "Fundamental rights protection, stakeholder engagement, democratic oversight",
        "eu_articles": "Article 27",
        "effort_hours": 80,
    },
    {
        "category": "Fundamental Rights",
        "functionality": "Enhanced Fundamental Rights Analysis",
        "description": "Detailed assessment of EU Charter rights: privacy (Art 7), data protection (Art 8), freedom of expression (Art 11), non-discrimination (Art 21), effective remedy (Art 47)",
        "code_automatable": "No",
        "ai_automatable": "Yes",
        "automation_notes": "AI can analyze system features against charter rights, identify risks, generate assessment reports. Legal expertise required for final review",
        "risk_addressed": "Human rights compliance, ethical AI, legal protection",
        "eu_articles": "Article 27, EU Charter of Fundamental Rights",
        "effort_hours": 40,
    },
    {
        "category": "Technical Documentation",
        "functionality": "Comprehensive Annex IV Documentation - Development Process",
        "description": "Detailed documentation of development methodology, design specifications, architecture, computational resources, lifecycle changes",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "AI can extract from code repositories, generate architecture diagrams, document APIs. Code analysis tools can automate much of this",
        "risk_addressed": "Technical transparency, reproducibility, audit capability",
        "eu_articles": "Article 11, Annex IV.3",
        "effort_hours": 80,
    },
    {
        "category": "Technical Documentation",
        "functionality": "Annex IV - Validation and Testing Documentation",
        "description": "Test strategies, test results, validation procedures, performance benchmarks, edge case analysis, stress testing results",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Test automation frameworks can generate this. AI can analyze test coverage, identify gaps, generate test reports automatically",
        "risk_addressed": "Quality assurance, reliability verification, safety validation",
        "eu_articles": "Article 11, Annex IV.6",
        "effort_hours": 100,
    },
    {
        "category": "Technical Documentation",
        "functionality": "Annex IV - Performance Metrics Documentation",
        "description": "Definition and documentation of accuracy metrics, performance targets, robustness specifications, limitations, acceptable performance ranges",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "Metrics collection can be automated. AI can analyze performance data, define baselines, identify performance patterns",
        "risk_addressed": "Performance transparency, expectations management, quality standards",
        "eu_articles": "Article 11, Annex IV.7",
        "effort_hours": 60,
    },
    {
        "category": "Technical Documentation",
        "functionality": "Annex IV - Risk Management Documentation",
        "description": "Risk management system description, risk register, risk assessments, mitigation strategies, monitoring procedures",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "AI can identify risks from system analysis, suggest mitigations, generate risk matrices. Risk evaluation requires human judgment",
        "risk_addressed": "Risk transparency, systematic risk management, mitigation tracking",
        "eu_articles": "Article 11, Annex IV.8",
        "effort_hours": 60,
    },
    {
        "category": "Technical Documentation",
        "functionality": "Annex IV - Cybersecurity Documentation",
        "description": "Cybersecurity specifications, threat model, security controls, penetration test results, vulnerability management",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "Security scanning tools can automate vulnerability detection. AI can analyze threats, suggest controls. Penetration testing requires security experts",
        "risk_addressed": "Security assurance, threat protection, vulnerability management",
        "eu_articles": "Article 11, Article 15, Annex IV.10",
        "effort_hours": 50,
    },
    {
        "category": "Logging and Traceability",
        "functionality": "Automatic Logging System (Article 12)",
        "description": "Privacy-preserving logging of interactions: query hashes, timestamps, response metadata, model versions, errors, compliance tracking",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Fully automatable with logging frameworks. AI can enhance with anomaly detection, pattern recognition, automated alerting",
        "risk_addressed": "Traceability, audit trail, incident investigation, accountability",
        "eu_articles": "Article 12",
        "effort_hours": 80,
    },
    {
        "category": "Accuracy and Performance",
        "functionality": "Accuracy Metrics and Testing",
        "description": "Define and measure factual accuracy (>95% correct citations), completeness (>90% complete answers), consistency (>85% consistency score)",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Fully automatable: automated testing, ground truth comparison, statistical analysis. AI can generate test cases, validate responses",
        "risk_addressed": "Output quality, reliability, user trust, regulatory compliance",
        "eu_articles": "Article 15.1, Article 15.3",
        "effort_hours": 120,
    },
    {
        "category": "Robustness Testing",
        "functionality": "Edge Case and Stress Testing",
        "description": "Testing with edge cases (very short/long queries), input perturbations (typos, formatting), boundary conditions, failure modes",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Automated testing frameworks can handle this. AI can generate edge cases, mutate inputs, identify failure patterns",
        "risk_addressed": "System robustness, graceful degradation, reliability",
        "eu_articles": "Article 15.3, Article 9.7",
        "effort_hours": 100,
    },
    {
        "category": "Robustness Testing",
        "functionality": "Adversarial Testing",
        "description": "Testing against adversarial inputs: prompt injection attempts, jailbreaking, manipulation attempts, security probing",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "Automated adversarial testing tools available. AI can generate adversarial examples, test defenses. Security expertise needed for comprehensive testing",
        "risk_addressed": "Security resilience, prompt injection protection, manipulation resistance",
        "eu_articles": "Article 15.3, Article 9.8",
        "effort_hours": 80,
    },
    {
        "category": "Cybersecurity",
        "functionality": "Comprehensive Cybersecurity Assessment",
        "description": "Penetration testing, threat modeling, security code review, secrets management, vulnerability assessment",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "Security scanning tools can automate vulnerability detection. AI can assist with threat modeling, code analysis. Penetration testing requires security experts",
        "risk_addressed": "System security, data protection, attack prevention",
        "eu_articles": "Article 15.1, Article 15.4",
        "effort_hours": 280,
    },
    {
        "category": "Cybersecurity",
        "functionality": "Input Validation and Sanitization",
        "description": "Validation of all user inputs, sanitization to prevent injection attacks, rate limiting, input filtering",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Fully automatable with input validation libraries. AI can learn attack patterns, enhance detection, adaptive filtering",
        "risk_addressed": "Injection attack prevention, system integrity, data protection",
        "eu_articles": "Article 15.4",
        "effort_hours": 40,
    },
    {
        "category": "Cybersecurity",
        "functionality": "Security Monitoring and Alerting",
        "description": "Real-time monitoring for security events, intrusion detection, automated alerting, security incident response",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Fully automatable with SIEM tools. AI excels at anomaly detection, threat prediction, automated response",
        "risk_addressed": "Threat detection, incident response, continuous security",
        "eu_articles": "Article 15.4, Article 73",
        "effort_hours": 60,
    },
    {
        "category": "Human Oversight",
        "functionality": "Enhanced Article 14 Controls",
        "description": "Explicit consent mechanism, enhanced verification steps, clear override controls, system understanding aids",
        "code_automatable": "Yes",
        "ai_automatable": "Partial",
        "automation_notes": "UI controls can be coded. AI can enhance with adaptive interfaces, personalized explanations. Human factors design needed",
        "risk_addressed": "Human control, autonomous decision prevention, user empowerment",
        "eu_articles": "Article 14",
        "effort_hours": 40,
    },
    {
        "category": "Transparency and Instructions",
        "functionality": "Enhanced Instructions for Use (Article 13)",
        "description": "Contact information, expected system lifetime, maintenance procedures, update process, support channels",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "Documentation can be generated with AI. Version management automatable. Human review required for accuracy",
        "risk_addressed": "User information, proper use, maintenance clarity, support access",
        "eu_articles": "Article 13",
        "effort_hours": 20,
    },
    {
        "category": "Incident Management",
        "functionality": "Serious Incident Reporting System (Article 73)",
        "description": "Incident detection, severity classification, 15-day reporting timeline tracking, remediation workflow, regulatory notification",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "Workflow automation possible. AI can detect incidents, classify severity, suggest remediation. Human judgment required for serious incidents",
        "risk_addressed": "Safety incidents, regulatory compliance, timely response",
        "eu_articles": "Article 73",
        "effort_hours": 60,
    },
    {
        "category": "Monitoring Infrastructure",
        "functionality": "Real-Time Monitoring Dashboard",
        "description": "Real-time performance dashboard showing accuracy, errors, usage, alerts, trends, system health indicators",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Fully automatable with dashboard tools (Grafana, etc.). AI can enhance with predictive analytics, intelligent alerting",
        "risk_addressed": "Operational visibility, proactive issue detection, performance tracking",
        "eu_articles": "Article 72, Article 15",
        "effort_hours": 100,
    },
    {
        "category": "Change Management",
        "functionality": "Systematic Change Management Process",
        "description": "Change request workflow, impact assessment, testing before deployment, rollback procedures, change documentation",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "Workflow automation possible. AI can assess change impact, predict risks, automate testing. Human approval required for changes",
        "risk_addressed": "Controlled updates, stability maintenance, regression prevention",
        "eu_articles": "Article 17, Article 43",
        "effort_hours": 60,
    },
    {
        "category": "Data Governance",
        "functionality": "Enhanced Data Governance Framework",
        "description": "Data quality monitoring, bias detection in training data, data lineage tracking, data retention policies, privacy controls",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Fully automatable: data quality checks, bias detection algorithms, lineage tracking tools. AI can enhance anomaly detection",
        "risk_addressed": "Data quality, bias prevention, privacy protection, regulatory compliance",
        "eu_articles": "Article 10",
        "effort_hours": 60,
    },
    {
        "category": "External Assurance",
        "functionality": "Third-Party Security Audit",
        "description": "Independent security assessment by external auditors, penetration testing, code review, vulnerability report",
        "code_automatable": "No",
        "ai_automatable": "Partial",
        "automation_notes": "AI can prepare audit materials, automate evidence collection. Actual audit requires human security experts",
        "risk_addressed": "Independent verification, security assurance, credibility",
        "eu_articles": "Article 15.4, Article 17",
        "effort_hours": 80,
    },
    {
        "category": "User Training",
        "functionality": "Deployer Training Materials",
        "description": "Training materials for deployers on proper use, limitations, oversight responsibilities, incident reporting",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "AI can generate training content, create interactive tutorials, personalize learning. Instructional design requires human expertise",
        "risk_addressed": "Proper use, deployer competence, oversight effectiveness",
        "eu_articles": "Article 13, Article 14",
        "effort_hours": 40,
    },
    {
        "category": "Bias and Fairness",
        "functionality": "Bias Testing and Monitoring",
        "description": "Regular bias testing across demographic groups, fairness metrics, bias detection in outputs, corrective actions",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Fully automatable with fairness testing libraries. AI can detect subtle biases, analyze patterns, suggest mitigations",
        "risk_addressed": "Non-discrimination, fairness, equitable treatment",
        "eu_articles": "Article 10, Article 27 (EU Charter Art 21)",
        "effort_hours": 60,
    },
    {
        "category": "Localization",
        "functionality": "Multilingual Support Documentation",
        "description": "Translation of documentation and instructions to EU languages, culturally appropriate content",
        "code_automatable": "Partial",
        "ai_automatable": "Yes",
        "automation_notes": "AI translation tools can automate most translation. Human review needed for technical accuracy and cultural appropriateness",
        "risk_addressed": "Accessibility, EU market coverage, user understanding",
        "eu_articles": "Article 13 (implicit - must be understandable)",
        "effort_hours": 80,
    },
    {
        "category": "Accessibility",
        "functionality": "Accessibility Compliance",
        "description": "WCAG compliance, screen reader support, keyboard navigation, accessible documentation",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Accessibility testing tools can automate compliance checking. AI can suggest improvements, test with simulated assistive technologies",
        "risk_addressed": "Disability inclusion, equal access, fundamental rights",
        "eu_articles": "Article 27 (EU Charter Art 26), General accessibility requirements",
        "effort_hours": 60,
    },
    {
        "category": "Environmental",
        "functionality": "Environmental Impact Optimization",
        "description": "Carbon footprint measurement, RAG optimization to reduce token usage, energy efficiency improvements",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Fully automatable: carbon tracking tools, optimization algorithms. AI can optimize queries, reduce computational waste",
        "risk_addressed": "Environmental sustainability, efficiency, responsible AI",
        "eu_articles": "Article 15 (efficiency implied), ESG compliance",
        "effort_hours": 40,
    },
    {
        "category": "Interoperability",
        "functionality": "Standards Compliance and Interoperability",
        "description": "Compliance with harmonized standards, API standardization, data format interoperability",
        "code_automatable": "Yes",
        "ai_automatable": "Yes",
        "automation_notes": "Standards compliance testing can be automated. AI can suggest standard-compliant implementations, validate conformance",
        "risk_addressed": "Interoperability, vendor lock-in prevention, ecosystem integration",
        "eu_articles": "Article 40 (harmonized standards)",
        "effort_hours": 60,
    }
]


def create_gap_analysis_sheet():
    """Create or update the Excel file with gap analysis sheet."""

    print(f"Opening Excel file: {EXCEL_PATH}")

    try:
        # Try to load existing workbook
        wb = openpyxl.load_workbook(EXCEL_PATH)
        print(f"Loaded existing workbook with sheets: {wb.sheetnames}")
    except FileNotFoundError:
        # Create new workbook if file doesn't exist
        wb = openpyxl.Workbook()
        print("Created new workbook")

    # Create new sheet or use existing
    sheet_name = "High-Risk Gap Analysis"
    if sheet_name in wb.sheetnames:
        print(f"Sheet '{sheet_name}' already exists. Removing it to create fresh...")
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name, 0)  # Insert at beginning

    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    category_font = Font(bold=True, size=10)

    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Define headers
    headers = [
        "Category",
        "Functionality Name",
        "Description",
        "Can Be Automated with Code?",
        "Can Be Coded with AI/LLM?",
        "Automation Notes",
        "Risk Type Addressed",
        "EU AI Act Article(s)",
        "Effort (Hours)"
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border_thin

    # Set column widths
    column_widths = {
        'A': 25,  # Category
        'B': 30,  # Functionality Name
        'C': 50,  # Description
        'D': 15,  # Can Be Automated with Code
        'E': 15,  # Can Be Coded with AI
        'F': 50,  # Automation Notes
        'G': 40,  # Risk Type Addressed
        'H': 20,  # EU AI Act Articles
        'I': 12,  # Effort Hours
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Write data
    current_row = 2
    current_category = None

    for item in MISSING_FUNCTIONALITY:
        # Track category changes for visual grouping
        if item['category'] != current_category:
            current_category = item['category']
            category_row = current_row

        # Write row data
        ws.cell(row=current_row, column=1).value = item['category']
        ws.cell(row=current_row, column=2).value = item['functionality']
        ws.cell(row=current_row, column=3).value = item['description']
        ws.cell(row=current_row, column=4).value = item['code_automatable']
        ws.cell(row=current_row, column=5).value = item['ai_automatable']
        ws.cell(row=current_row, column=6).value = item['automation_notes']
        ws.cell(row=current_row, column=7).value = item['risk_addressed']
        ws.cell(row=current_row, column=8).value = item['eu_articles']
        ws.cell(row=current_row, column=9).value = item['effort_hours']

        # Apply formatting
        for col_num in range(1, 10):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = border_thin
            cell.alignment = wrap_alignment

            # Center align certain columns
            if col_num in [4, 5, 9]:
                cell.alignment = Alignment(horizontal='center', vertical='top')

        current_row += 1

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add summary sheet
    create_summary_sheet(wb, MISSING_FUNCTIONALITY)

    # Save workbook
    wb.save(EXCEL_PATH)
    print(f"\n✅ Excel file updated successfully: {EXCEL_PATH}")
    print(f"   - Sheet '{sheet_name}' created with {len(MISSING_FUNCTIONALITY)} missing functionalities")

    return wb


def create_summary_sheet(wb, data):
    """Create a summary analysis sheet."""

    sheet_name = "Gap Analysis Summary"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name, 1)

    # Calculate summaries
    total_items = len(data)
    total_hours = sum(item['effort_hours'] for item in data)

    # Count by automation potential
    fully_automatable_code = sum(1 for item in data if item['code_automatable'] == 'Yes')
    partially_automatable_code = sum(1 for item in data if item['code_automatable'] == 'Partial')
    not_automatable_code = sum(1 for item in data if item['code_automatable'] == 'No')

    fully_automatable_ai = sum(1 for item in data if item['ai_automatable'] == 'Yes')
    partially_automatable_ai = sum(1 for item in data if item['ai_automatable'] == 'Partial')
    not_automatable_ai = sum(1 for item in data if item['ai_automatable'] == 'No')

    # Count by category
    categories = {}
    for item in data:
        cat = item['category']
        if cat not in categories:
            categories[cat] = {'count': 0, 'hours': 0}
        categories[cat]['count'] += 1
        categories[cat]['hours'] += item['effort_hours']

    # Write summary
    row = 1
    ws.cell(row=row, column=1).value = "HIGH-RISK AI SYSTEM - GAP ANALYSIS SUMMARY"
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 1

    ws.cell(row=row, column=1).value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    row += 2

    # Overall summary
    ws.cell(row=row, column=1).value = "OVERALL SUMMARY"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    ws.cell(row=row, column=1).value = "Total Missing Functionalities:"
    ws.cell(row=row, column=2).value = total_items
    row += 1

    ws.cell(row=row, column=1).value = "Total Estimated Effort:"
    ws.cell(row=row, column=2).value = f"{total_hours} hours"
    row += 2

    # Automation potential - Code
    ws.cell(row=row, column=1).value = "AUTOMATION POTENTIAL - CODE"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    ws.cell(row=row, column=1).value = "Fully Automatable with Code:"
    ws.cell(row=row, column=2).value = f"{fully_automatable_code} ({fully_automatable_code/total_items*100:.1f}%)"
    row += 1

    ws.cell(row=row, column=1).value = "Partially Automatable with Code:"
    ws.cell(row=row, column=2).value = f"{partially_automatable_code} ({partially_automatable_code/total_items*100:.1f}%)"
    row += 1

    ws.cell(row=row, column=1).value = "Not Automatable with Code:"
    ws.cell(row=row, column=2).value = f"{not_automatable_code} ({not_automatable_code/total_items*100:.1f}%)"
    row += 2

    # Automation potential - AI
    ws.cell(row=row, column=1).value = "AUTOMATION POTENTIAL - AI/LLM"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    ws.cell(row=row, column=1).value = "Fully Automatable with AI:"
    ws.cell(row=row, column=2).value = f"{fully_automatable_ai} ({fully_automatable_ai/total_items*100:.1f}%)"
    row += 1

    ws.cell(row=row, column=1).value = "Partially Automatable with AI:"
    ws.cell(row=row, column=2).value = f"{partially_automatable_ai} ({partially_automatable_ai/total_items*100:.1f}%)"
    row += 1

    ws.cell(row=row, column=1).value = "Not Automatable with AI:"
    ws.cell(row=row, column=2).value = f"{not_automatable_ai} ({not_automatable_ai/total_items*100:.1f}%)"
    row += 2

    # Category breakdown
    ws.cell(row=row, column=1).value = "BREAKDOWN BY CATEGORY"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    ws.cell(row=row, column=1).value = "Category"
    ws.cell(row=row, column=2).value = "Count"
    ws.cell(row=row, column=3).value = "Hours"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 1

    for cat, stats in sorted(categories.items()):
        ws.cell(row=row, column=1).value = cat
        ws.cell(row=row, column=2).value = stats['count']
        ws.cell(row=row, column=3).value = stats['hours']
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    print(f"   - Summary sheet created")


def main():
    """Main execution function."""
    print("=" * 80)
    print("HIGH-RISK AI SYSTEM GAP ANALYSIS")
    print("Analyzing missing functionality for EU AI Act compliance")
    print("=" * 80)
    print()

    wb = create_gap_analysis_sheet()

    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Total missing functionalities: {len(MISSING_FUNCTIONALITY)}")
    print(f"Total estimated effort: {sum(item['effort_hours'] for item in MISSING_FUNCTIONALITY)} hours")
    print()
    print(f"Fully automatable with code: {sum(1 for item in MISSING_FUNCTIONALITY if item['code_automatable'] == 'Yes')}")
    print(f"Fully automatable with AI: {sum(1 for item in MISSING_FUNCTIONALITY if item['ai_automatable'] == 'Yes')}")
    print("=" * 80)


if __name__ == "__main__":
    main()
